            #Importerar frågor och facit, 5 st, från ett excel dokument
from openpyxl import load_workbook
workbook = load_workbook(filename="/Users/andreaslindblad/Desktop/BackUpCloud/quizdb.xlsx")

sheet = workbook.active

cell_value1 = sheet.cell(row=2, column=1).value
cell_value2 = sheet.cell(row=3, column=1).value
cell_value3 = sheet.cell(row=4, column=1).value
cell_value4 = sheet.cell(row=5, column=1).value
cell_value5 = sheet.cell(row=6, column=1).value

celll_facA1 = sheet.cell(row=2, column=2)
celll_facA2 = sheet.cell(row=3, column=2)
celll_facA3 = sheet.cell(row=4, column=2)
celll_facA4 = sheet.cell(row=5, column=2)
celll_facA5 = sheet.cell(row=6, column=2)
            #Slut excel import

end = False
while end == False:

    print(cell_value1)
    ans1 =  int(input("Svar: "))
    print(cell_value2)
    ans2 = int(input("Svar: "))
    print(cell_value3)
    ans3 = int(input("Svar: "))
    print(cell_value4)
    ans4 = int(input("Svar: "))
    print(cell_value5)
    ans5 = int(input("Svar:"))
            #Variablar för poäng
    result = 0
    fel = 0
            #Slut
            #Variabler för correction
    cor1 = False
    cor2 = False
    cor3 = False
    cor4 = False
    cor5 = False
            #Slut
            #Variabel som kontroll av grade
    gra = False
            #Slut
    if ans1 == celll_facA1.value:
        result = result + 1
    else:
        fel = fel + 1
        if ans1 != celll_facA1.value:
            cor1 = True
        else:
            cor1 = False

    if ans2 ==  celll_facA2.value:
        result = result + 1
    else:
        fel = fel + 1
        if ans2 != celll_facA2.value:
            cor2 = True
        else:
            cor2 = False
    if ans3 == celll_facA3.value:
        result = result + 1
    else:
        fel = fel + 1
        if ans3 != celll_facA3.value:
            cor3 = True
        else:
            cor3 = False
    if ans4 == celll_facA4.value:
        result = result + 1
    else:
        fel = fel + 1
        if ans4 != celll_facA4.value:
            cor4 = True
        else:
            cor4 = False
    if ans5 == celll_facA5.value:
        result = result +1
    else:
        fel = fel + 1
        if ans5 != celll_facA5.value:
            cor5 = True
        else:
            cor5 = False

    if result >= 3:
        gra = True
    else:
        pass

    print(" Facit  " + str(result) + "/5" )
    if gra == True:
        print ("Du är godkänt ")
    else:
        print("Du är ICKE godkänd")
        pass
    if result == 5:
        print("Grattis alla rätt")
    else:
        pass
    if cor1 == True:
        print("Du hade fel på Fråga 1")
        print("Rätt svar är " + str(celll_facA1.value))
    else:
        pass
    if cor2 == True:
        print("Du hade fel på Fråga 2")
        print("Rätt svar är " + str(celll_facA2.value))
    else:
        pass
    if cor3 == True:
        print("Du hade fel på Fråga 3")
        print("Rätt svar är " + str(celll_facA3.value))
    else:
        pass
    if cor4 == True:
        print("Du hade fel på Fråga 4")
        print("Rätt svar är " + str(celll_facA4.value))
    else:
        pass
    if cor5 == True:
        print("Du hade fel på Fråga 5")
        print("Rätt svar är " + str(celll_facA5.value))
    else:
        pass

    if gra == False:
        endt = input("Försöka igen? Y/N: ")
        if endt.upper ()== "N":
            end = True
        else:
            pass

    else:
        end = True




